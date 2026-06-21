import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Excelファイルを読み込む
df_2022 = pd.read_excel("2022_年間売上表.xlsx")
df_2023 = pd.read_excel("2023_年間売上表.xlsx")

# データを結合
df = pd.concat([df_2022, df_2023])

# 商品・売上年ごとに集計
summary = df.groupby(
    ["商品", "売上年"],
    as_index=False
)["金額（千円）"].sum()

# Excelへ出力
summary.to_excel(
    "売上集計表.xlsx",
    index=False
)

# 書式設定
wb = load_workbook("売上集計表.xlsx")
ws = wb.active

gray_fill = PatternFill(
    fill_type="solid",
    fgColor="F2F2F2"
)

for cell in ws[1]:
    cell.fill = gray_fill

wb.save("売上集計表.xlsx")

print("売上集計表.xlsxを作成しました")