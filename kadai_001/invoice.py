import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "請求書"

today = datetime.today().strftime("%Y%m%d")
display_date = datetime.today().strftime("%Y/%m/%d")

file_name = f"請求書_{today}.xlsx"

ws["B2"] = "請求書"
ws["B2"].font = Font(size=20, bold=True)

ws["B4"] = "株式会社ABC"
ws["B5"] = "〒101-0022 東京都千代田区神田練塀町300"
ws["B6"] = "TEL:03-1234-5678 FAX:03-1234-5678"
ws["B7"] = "担当者名:鈴木一郎 様"

ws["F4"] = "No."
ws["G4"] = "0001"

ws["F5"] = "日付"
ws["G5"] = display_date

ws["B10"] = "商品名"
ws["C10"] = "数量"
ws["D10"] = "単価"
ws["E10"] = "金額"

items = [
    ["商品A", 2, 10000],
    ["商品B", 1, 15000],
]

start_row = 11

for i, item in enumerate(items):
    row = start_row + i
    name = item[0]
    quantity = item[1]
    price = item[2]
    amount = quantity * price

    ws.cell(row=row, column=2, value=name)
    ws.cell(row=row, column=3, value=quantity)
    ws.cell(row=row, column=4, value=price)
    ws.cell(row=row, column=5, value=amount)

total_row = start_row + len(items) + 1

subtotal = f"=SUM(E{start_row}:E{start_row + len(items) - 1})"
tax = f"=E{total_row}*0.1"
total = f"=E{total_row}+E{total_row + 1}"

ws.cell(row=total_row, column=2, value="小計")
ws.cell(row=total_row, column=5, value=subtotal)

ws.cell(row=total_row + 1, column=2, value="消費税")
ws.cell(row=total_row + 1, column=5, value=tax)

ws.cell(row=total_row + 2, column=2, value="合計")
ws.cell(row=total_row + 2, column=5, value=total)

thin = Side(style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

for row in ws["B10:E17"]:
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 10
ws.column_dimensions["G"].width = 15

wb.save(file_name)
print(file_name)